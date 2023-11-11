import openpyxl


workbook = openpyxl.Workbook()
sheet = workbook.active

sheet["A1"] = "Mega Carros"
sheet["A2"] = "Salir"
sheet["B2"] = "Codigo"
sheet["C2"] = "Marca"
sheet["D2"]= "Modelo"
sheet["E2"] = "Precio"
sheet["F2"] = "Kilometraje"

sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

siguiente_fila = sheet.max_row + 1



def guardar():
    workbook = openpyxl.Workbook()
sheet = workbook.active

sheet["A1"] = "Mega Carros"
sheet["A2"] = "Salir"
sheet["B2"] = "Codigo"
sheet["C2"] = "Marca"
sheet["D2"]= "Modelo"
sheet["E2"] = "Precio"
sheet["F2"] = "Kilometraje"

sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

siguiente_fila = sheet.max_row + 1

def guardar():
    with open('vehiculos.xlsx', 'w') as archivo:
        for vehiculo in vehiculos:
            archivo.write(f"{vehiculo.codigo},{vehiculo.marca},{vehiculo.modelo}\n")   
 
def crear(): 
    while True:
      salida = input("si desea salir, escriba 'salir': ")
      if not salida: #al presionar enter se sale
            break
      codigo = float(input("Codigo del vehiculo: "))
      marca =  input("Marxa del vehiculo: ")
      modelo = input("Modelo del vehiculo: ")
      precio = float(input("Precio del vechiculo: "))
      kilometraje =   int(input("Kilometraje del vehiculo: "))
    
    
    sheet[f"A{fila}"]= salida
    sheet[f"B{fila}"]= codigo
    sheet[f"C{fila}"]= marca
    sheet[f"D{fila}"]= modelo
    sheet[f"E{fila}"]= precio
    sheet[f"F{fila}"]= kilometraje

fila =+ 1

workbook.save("vehiculos.xlsx")

def editar():
    codigo = input("Ingrese el código del auto")

    for vehiculo in vehiculos:
        if vehiculo.codigo == codigo:
            print("Editar vehiculo:")
            nueva_marca = input(f"marca: {vehiculo.marca}. Nueva marca: ")
            nuevo_precio = input(f"Precio actual: {vehiculo.precio}. Nueo precio: ")

            if nueva_marca:
                vehiculo.marca = nueva_marca
            if nuevo_precio:
                vehiculo.precio = nuevo_precio

            print("auto actuyalizado")
            vehiculo = Vehiculo(codigo, marca, modelo)
            vehiculo.append(vehiculo)
            workbook.save("vehiculos.xlsx")
        
    
    
def eliminar():
    codigo = input("Ingrese el código del cliente que desea eliminar: ")

    for vehiculos in vehiculo:
        if Vehiculo.codigo == codigo:
            vehiculo.remove(vehiculo)
            print(f"vehiculo '{vehiculo.marca}' eliminado exitosamente.")
            return
    
    print(f"No se encontró ningún cliente con el código '{codigo}'.")
    
    
    
def listar(): 
    if not vehiculos:
        print("No hay vehiculos")
        return
    
    print("Lista de clientes:")
    for vehiculo in vehiculos:
        print(f"Código: {vehiculo.codigo}, marca: {vehiculo.marca}, modelos: {vehiculo.modelo}")


while True:
    opcion = input("Elije: ")
    if opcion.lower() == 'salir':
        break
    

    if opcion == '1':
        crear()
    elif opcion == '2':
        editar()
    elif opcion == '3':
        eliminar()
    elif opcion == '4':
        listar()
    
    
    
    

  